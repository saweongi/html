#pip install openpyxl
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

workbook = openpyxl.load_workbook('test.xlsx') #test라는 엑셀을 열어준다.
worksheet = workbook.active # "활성화되어있는 Sheet"에 접근하는 방법
worksheet.merge_cells('D1:E1') # 셀 병합(d1열이랑 e1열을 병합한다.)
worksheet['A1'].font = Font(bold =True, size=24) #폰트를 굵고 사이즈를 24로 키운다
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') #처음과 끝의 색깔을 FFFF00으로 하고  fill_type 은 solid라는 배경색으로 한다.
worksheet['A1'].fill = yellow_fill
worksheet.column_dimensions['A'].width = 100 #열의 너비를 설정
worksheet.row_dimensions[1].height = 100 #1행 높이 를 설정
for row in worksheet.iter_rows(values_only=True): #하나의 행을 가져오게한다
    print(row)
workbook.close()
workbook.save('test2.xlsx')